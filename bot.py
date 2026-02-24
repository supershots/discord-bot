import discord
import anthropic
import os
from googleapiclient.discovery import build
from google.oauth2 import service_account
import io
from googleapiclient.http import MediaIoBaseDownload
import docx
import openpyxl

DISCORD_TOKEN = os.environ.get("DISCORD_TOKEN")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY")
FOLDER_ID = "1Weoz06pU3Dxjw5kRA8fq5BKX5N8FOFAn"

def load_drive_files():
    import json
creds_json = os.environ.get("GOOGLE_CREDENTIALS")
creds_dict = json.loads(creds_json)
creds = service_account.Credentials.from_service_account_info(
    creds_dict,
    scopes=["https://www.googleapis.com/auth/drive.readonly"]
)
    service = build("drive", "v3", credentials=creds)
    results = service.files().list(
        q=f"'{FOLDER_ID}' in parents",
        fields="files(id, name, mimeType)"
    ).execute()
    files = results.get("files", [])
    all_text = ""
    for file in files:
        try:
            if file["mimeType"] == "application/vnd.google-apps.document":
                # Googleドキュメント
                content = service.files().export(
                    fileId=file["id"], mimeType="text/plain"
                ).execute()
                all_text += f"\n【{file['name']}】\n{content.decode('utf-8')}\n"
            elif "wordprocessingml" in file["mimeType"]:
                # Wordファイル
                request = service.files().get_media(fileId=file["id"])
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                fh.seek(0)
                doc = docx.Document(fh)
                text = "\n".join([p.text for p in doc.paragraphs])
                all_text += f"\n【{file['name']}】\n{text}\n"
            elif "spreadsheetml" in file["mimeType"]:
                # Excelファイル
                request = service.files().get_media(fileId=file["id"])
                fh = io.BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
                fh.seek(0)
                wb = openpyxl.load_workbook(fh)
                text = ""
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows(values_only=True):
                        text += " ".join([str(c) for c in row if c]) + "\n"
                all_text += f"\n【{file['name']}】\n{text}\n"
        except Exception as e:
            print(f"{file['name']} の読み込みエラー：{e}")
    return all_text

intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)
ai_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

print("Google Driveからファイルを読み込み中...")
knowledge = load_drive_files()
print(f"読み込み完了！{len(knowledge)}文字分の情報を取得しました")

SYSTEM_PROMPT = f"""
あなたは就労継続支援の作業をサポートするアシスタントです。
利用者の方が作業中に困ったことを質問するので、
わかりやすく、やさしい言葉で答えてください。
以下の情報をもとに答えてください。

{knowledge}
"""

@client.event
async def on_ready():
    print(f"Botが起動しました: {client.user}")

@client.event
async def on_message(message):
    if message.author == client.user:
        return
    if message.channel.name != "作業の質問":
        return
    thinking_msg = await message.reply("考え中です...少し待ってね🤔")
    try:
        response = ai_client.messages.create(
            model="claude-opus-4-6",
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": message.content}]
        )
        answer = response.content[0].text
        await thinking_msg.edit(content=answer)
    except Exception as e:
        await thinking_msg.edit(content=f"エラー：{e}")

client.run(DISCORD_TOKEN)